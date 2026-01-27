"""
Create Test Results Excel
创建测试结果记录表格
"""
import sys
sys.path.insert(0, 'C:/test/antinet/backend')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# 创建工作簿
wb = Workbook()

# 工作表 1：NPU 性能测试结果
ws1 = wb.active
ws1.title = "NPU Performance"

# 标题行
headers = ["测试时间", "测试项目", "指标", "实际值", "目标值", "状态", "备注"]
ws1.append(headers)

# 设置标题样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for cell in ws1[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 添加示例数据
test_data = [
    [datetime.now().strftime("%Y-%m-%d %H:%M"), "模型加载", "加载时间", "待测试", "10-15秒", "待测试", ""],
    [datetime.now().strftime("%Y-%m-%d %H:%M"), "短文本推理", "平均延迟", "待测试", "< 500ms", "待测试", ""],
    [datetime.now().strftime("%Y-%m-%d %H:%M"), "长文本推理", "延迟", "待测试", "< 2000ms", "待测试", ""],
]

for row in test_data:
    ws1.append(row)

# 调整列宽
ws1.column_dimensions['A'].width = 18
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 15
ws1.column_dimensions['E'].width = 15
ws1.column_dimensions['F'].width = 12
ws1.column_dimensions['G'].width = 30

# 工作表 2：API 测试结果
ws2 = wb.create_sheet("API Tests")
ws2.append(["测试时间", "API 端点", "状态码", "响应时间(ms)", "状态", "备注"])

# 设置标题样式
for cell in ws2[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

api_tests = [
    [datetime.now().strftime("%Y-%m-%d %H:%M"), "/api/health", "待测试", "待测试", "待测试", ""],
    [datetime.now().strftime("%Y-%m-%d %H:%M"), "/api/skill/list", "待测试", "待测试", "待测试", ""],
    [datetime.now().strftime("%Y-%m-%d %H:%M"), "/api/knowledge/graph", "待测试", "待测试", "待测试", ""],
]

for row in api_tests:
    ws2.append(row)

# 调整列宽
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws2.column_dimensions[col].width = 18

# 工作表 3：功能测试清单
ws3 = wb.create_sheet("Feature Checklist")
ws3.append(["功能模块", "测试项目", "状态", "测试时间", "测试人", "备注"])

# 设置标题样式
for cell in ws3[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

features = [
    ["知识图谱", "图谱可视化", "待测试", "", "", ""],
    ["知识图谱", "节点拖拽", "待测试", "", "", ""],
    ["知识图谱", "边关系显示", "待测试", "", "", ""],
    ["技能系统", "24个技能注册", "待测试", "", "", ""],
    ["技能系统", "技能执行", "待测试", "", "", ""],
    ["数据分析", "四色卡片生成", "待测试", "", "", ""],
    ["NPU 推理", "模型加载", "待测试", "", "", ""],
    ["NPU 推理", "推理延迟", "待测试", "", "", ""],
]

for row in features:
    ws3.append(row)

# 调整列宽
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws3.column_dimensions[col].width = 18

# 保存文件
output_file = "C:/test/antinet/test_results.xlsx"
wb.save(output_file)

print(f"Test results template created: {output_file}")
print("You can open it in Excel and update the test results manually.")
