"""
Excel Export API Routes
提供 Excel 导出功能的 API 端点
"""

from fastapi import APIRouter, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
import os
from datetime import datetime
from pathlib import Path
import pandas as pd

# 注释掉不存在的导入
# from skills.xlsx import export_cards_to_excel, export_analysis_to_excel

router = APIRouter(prefix="/api/excel", tags=["excel"])

OUTPUT_DIR = Path("./data/exports")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


@router.get("/health")
async def health_check():
    """Excel 服务健康检查"""
    return {
        "status": "healthy",
        "service": "excel",
        "available": True
    }


class CardExportRequest(BaseModel):
    """卡片导出请求"""
    cards: List[Dict[str, Any]]
    title: Optional[str] = "Antinet 卡片导出"
    filename: Optional[str] = None


class AnalysisExportRequest(BaseModel):
    """分析报告导出请求"""
    analysis_info: Dict[str, Any]
    cards_by_type: Dict[str, List[Dict[str, Any]]]
    data_sheets: Optional[Dict[str, Any]] = None
    charts: Optional[List[Dict[str, Any]]] = None
    filename: Optional[str] = None


@router.post("/export-cards")
async def export_cards(request: CardExportRequest):
    """
    导出卡片到 Excel
    
    请求体示例：
    ```json
    {
        "cards": [
            {
                "id": "fact_001",
                "type": "fact",
                "title": "销售数据",
                "content": "2025年1月销售额为100万",
                "confidence": 0.95,
                "created_at": "2025-01-26",
                "tags": ["销售", "数据"]
            }
        ],
        "title": "销售分析卡片",
        "filename": "sales_cards.xlsx"
    }
    ```
    """
    try:
        # 生成文件名
        if request.filename:
            filename = request.filename
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"cards_export_{timestamp}.xlsx"
        
        output_path = OUTPUT_DIR / filename

        # 简单导出（使用pandas）
        cards_df = pd.DataFrame(request.cards)
        cards_df.to_excel(output_path, index=False, engine='openpyxl')

        return {
            "status": "success",
            "message": "卡片导出成功",
            "filename": filename,
            "path": str(output_path),
            "download_url": f"/api/excel/download/{filename}"
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


def _generate_four_color_cards_from_data(data: List[Dict], stats: Dict) -> List[Dict]:
    """从Excel数据生成4色卡片"""
    cards = []
    
    # 生成4色卡片
    # 1. 🔵 事实卡片 - 描述数据基本事实
    fact_card = {
        '卡片ID': 1,
        '类型': '🔵 事实卡片',
        '标题': '数据概览',
        '内容': f"本次分析共处理 {stats.get('totalRows', 0)} 行数据，包含 {stats.get('totalColumns', 0)} 个列。其中数值列 {stats.get('numericColumns', 0)} 个，文本列 {stats.get('textColumns', 0)} 个。",
        '置信度': 1.0,
        '创建时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        '标签': '数据分析,统计'
    }
    cards.append(fact_card)
    
    # 2. 🟢 解释卡片 - 解释数据含义
    interpret_card = {
        '卡片ID': 2,
        '类型': '🟢 解释卡片',
        '标题': '数据质量分析',
        '内容': f"数据完整性分析：发现 {stats.get('missingValues', 0)} 个缺失值，缺失率约为 {stats.get('missingValues', 0) / (stats.get('totalRows', 1) * stats.get('totalColumns', 1)) * 100:.2f}%。重复数据 {stats.get('duplicates', 0)} 行。",
        '置信度': 0.9,
        '创建时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        '标签': '数据质量,解读'
    }
    cards.append(interpret_card)
    
    # 3. 🟡 风险卡片 - 识别潜在风险
    risk_card = {
        '卡片ID': 3,
        '类型': '🟡 风险卡片',
        '标题': '数据风险提示',
        '内容': f"风险提示：{'数据存在较多缺失值，可能影响分析准确性。' if stats.get('missingValues', 0) > 0 else '数据质量良好，无明显风险。'}{'存在重复数据，建议清理。' if stats.get('duplicates', 0) > 0 else ''}",
        '置信度': 0.8,
        '创建时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        '标签': '风险提示,质量'
    }
    cards.append(risk_card)
    
    # 4. 🔴 行动建议 - 提供改进建议
    action_card = {
        '卡片ID': 4,
        '类型': '🔴 行动建议',
        '标题': '优化建议',
        '内容': f"建议：{'清理缺失值或补充数据。' if stats.get('missingValues', 0) > 0 else ''}{'删除重复数据。' if stats.get('duplicates', 0) > 0 else ''}继续进行深入的数据探索和可视化分析。",
        '置信度': 0.85,
        '创建时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        '标签': '行动建议,优化'
    }
    cards.append(action_card)
    
    return cards


@router.post("/export-analysis-simple")
async def export_analysis_simple(request: AnalysisExportRequest):
    """
    导出Excel分析为4色卡片格式
    
    请求体示例：
    ```json
    {
        "data": [{"列1": "值1", "列2": "值2"}],
        "columns": [{"key": "列1", "name": "列1", "type": "string"}],
        "stats": {"totalRows": 10, "totalColumns": 2},
        "filename": "export.xlsx"
    }
    ```
    """
    try:
        # 生成文件名
        if request.filename:
            filename = request.filename
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"analysis_cards_{timestamp}.xlsx"
        
        output_path = OUTPUT_DIR / filename
        
        # 从请求中获取数据
        data = getattr(request, 'data', []) if hasattr(request, 'data') else []
        stats = getattr(request, 'stats', {}) if hasattr(request, 'stats') else {}
        
        # 生成4色卡片
        cards = _generate_four_color_cards_from_data(data, stats)
        
        # 创建卡片DataFrame
        df_cards = pd.DataFrame(cards)
        
        # 导出到Excel
        df_cards.to_excel(output_path, index=False, engine='openpyxl')
        
        # 应用4色标记
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 根据卡片类型设置行颜色
        type_colors = {
            '🔵 事实卡片': '5B9BD5',   # 蓝色
            '🟢 解释卡片': '70AD47',   # 绿色
            '🟡 风险卡片': 'FFC000',   # 黄色
            '🔴 行动建议': 'ED7D31'    # 橙红色
        }
        
        # 表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
        
        # 根据类型设置行颜色
        for row_idx in range(2, ws.max_row + 1):
            card_type = ws.cell(row=row_idx, column=2).value  # 类型列
            if card_type in type_colors:
                fill_color = type_colors[card_type]
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color=fill_color,
                        end_color=fill_color,
                        fill_type='solid'
                    )
        
        wb.save(output_path)
        
        return {
            "status": "success",
            "message": "数据导出成功",
            "filename": filename,
            "path": str(output_path),
            "download_url": f"/api/excel/download/{filename}"
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.post("/export-analysis")
async def export_analysis(request: AnalysisExportRequest):
    """
    导出完整分析报告到 Excel
    
    请求体示例：
    ```json
    {
        "analysis_info": {
            "title": "2025年1月销售分析报告",
            "date": "2025-01-26",
            "data_source": "sales_data.csv",
            "card_counts": {
                "fact": 5,
                "interpret": 3,
                "risk": 2,
                "action": 4
            },
            "summary": "本报告分析了1月销售数据..."
        },
        "cards_by_type": {
            "fact": [...],
            "interpret": [...],
            "risk": [...],
            "action": [...]
        },
        "filename": "sales_analysis_report.xlsx"
    }
    ```
    """
    try:
        # 生成文件名
        if request.filename:
            filename = request.filename
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"analysis_report_{timestamp}.xlsx"
        
        output_path = OUTPUT_DIR / filename
        
        # 转换 data_sheets (如果有)
        import pandas as pd
        data_sheets = None
        if request.data_sheets:
            data_sheets = {}
            for sheet_name, sheet_data in request.data_sheets.items():
                data_sheets[sheet_name] = pd.DataFrame(sheet_data)

        # 简单导出
        # 这里可以添加更多导出逻辑
        all_cards = []
        for card_type, cards in request.cards_by_type.items():
            for card in cards:
                card['type'] = card_type
                all_cards.append(card)

        cards_df = pd.DataFrame(all_cards)
        cards_df.to_excel(output_path, index=False, engine='openpyxl')

        return {
            "status": "success",
            "message": "分析报告导出成功",
            "filename": filename,
            "path": str(output_path),
            "download_url": f"/api/excel/download/{filename}"
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.get("/download/{filename}")
async def download_file(filename: str, background_tasks: BackgroundTasks):
    """
    下载导出的 Excel 文件
    
    Args:
        filename: 文件名
    """
    file_path = OUTPUT_DIR / filename
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="文件不存在")
    
    # 添加后台任务：下载后删除文件（可选）
    # background_tasks.add_task(os.remove, file_path)
    
    return FileResponse(
        path=str(file_path),
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.get("/list")
async def list_exports():
    """
    列出所有导出的文件
    """
    try:
        files = []
        for file_path in OUTPUT_DIR.glob("*.xlsx"):
            stat = file_path.stat()
            files.append({
                "filename": file_path.name,
                "size": stat.st_size,
                "created_at": datetime.fromtimestamp(stat.st_ctime).isoformat(),
                "download_url": f"/api/excel/download/{file_path.name}"
            })
        
        return {
            "status": "success",
            "count": len(files),
            "files": sorted(files, key=lambda x: x['created_at'], reverse=True)
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取文件列表失败: {str(e)}")


@router.delete("/delete/{filename}")
async def delete_file(filename: str):
    """
    删除导出的文件
    
    Args:
        filename: 文件名
    """
    file_path = OUTPUT_DIR / filename
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="文件不存在")
    
    try:
        os.remove(file_path)
        return {
            "status": "success",
            "message": f"文件 {filename} 已删除"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"删除失败: {str(e)}")
