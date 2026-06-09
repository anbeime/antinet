"""
Invoice API Routes — 发票处理与查询

命令适配（意图识别）：
  - 扫描发票 -> scan
  - 列出发票 -> list
  - 查询商家 -> query --seller XX
  - 导出报销 -> export --from ... --to ...
  - 排除发票 -> exclude <id>
  - 恢复发票 -> include <id>
  - 查看问题 -> problems
"""
import logging
import os
import re
import sys
import json
import shutil
import zipfile
import tempfile
import io
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Query, BackgroundTasks, Request
from fastapi.responses import FileResponse
from pydantic import BaseModel

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/invoice", tags=["发票管理"])

INVOICE_DIR = Path("./data/invoices")
INVOICE_DIR.mkdir(parents=True, exist_ok=True)

OUTPUT_DIR = Path("./data/exports")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

db_manager = None
_migrated = False


def set_db_manager(dbm):
    global db_manager, _migrated
    db_manager = dbm
    if not _migrated:
        _migrated = True
        _migrate_existing_data()


def _migrate_existing_data():
    """Fix existing records where amounts are stored as strings with currency symbols"""
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        for col in ("total_amount", "amount", "tax_amount"):
            cursor.execute(f"SELECT id, {col} FROM invoices WHERE {col} IS NOT NULL")
            for row in cursor.fetchall():
                val = row[col]
                if isinstance(val, str):
                    cleaned = _clean_amount(val)
                    if cleaned is not None:
                        cursor.execute(f"UPDATE invoices SET {col}=?, updated_at=? WHERE id=?",
                                       (cleaned, datetime.now().isoformat(), row["id"]))
        conn.commit()
        conn.close()
        logger.info(f"[Invoice] Data migration completed")
    except Exception as e:
        logger.warning(f"[Invoice] Data migration skipped: {e}")


# ---- Models ----

class InvoiceInfo(BaseModel):
    id: int
    filename: str
    invoice_number: Optional[str] = None
    invoice_code: Optional[str] = None
    invoice_date: Optional[str] = None
    seller_name: Optional[str] = None
    seller_tax_id: Optional[str] = None
    buyer_name: Optional[str] = None
    buyer_tax_id: Optional[str] = None
    total_amount: Optional[float] = None
    amount: Optional[float] = None
    tax_amount: Optional[float] = None
    is_excluded: bool = False
    status: str = "pending"
    engine_used: Optional[str] = None
    created_at: str
    file_size: Optional[int] = None
    has_source_file: bool = False
    source_url: Optional[str] = None


def _safe_filename(name: str, fallback: str = "invoice") -> str:
    """Sanitize a filename for use in download Content-Disposition."""
    name = (name or fallback).strip()
    name = re.sub(r"[\\/:*?\"<>|\r\n\t]", "_", name)
    return name[:120] or fallback


class InvoiceListResponse(BaseModel):
    status: str
    count: int
    invoices: List[InvoiceInfo]


class InvoiceQueryParams(BaseModel):
    seller: Optional[str] = None
    buyer: Optional[str] = None
    from_date: Optional[str] = None
    to_date: Optional[str] = None
    status: Optional[str] = None


# ---- Helpers ----

def _clean_amount(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        v = v.strip().replace(",", "").replace("¥", "").replace("￥", "").replace("元", "")
        try:
            return float(v)
        except ValueError:
            return None
    return None


def _row_to_invoice(row, source_url_builder=None) -> InvoiceInfo:
    file_path = row["file_path"]
    has_source = bool(file_path) and os.path.exists(file_path)
    file_size = row["file_size"] or (os.path.getsize(file_path) if has_source else None)
    source_url = None
    if source_url_builder and has_source:
        source_url = source_url_builder(row["id"])
    return InvoiceInfo(
        id=row["id"],
        filename=row["filename"],
        invoice_number=row["invoice_number"],
        invoice_code=row["invoice_code"],
        invoice_date=row["invoice_date"],
        seller_name=row["seller_name"],
        seller_tax_id=row["seller_tax_id"],
        buyer_name=row["buyer_name"],
        buyer_tax_id=row["buyer_tax_id"],
        total_amount=_clean_amount(row["total_amount"]),
        amount=_clean_amount(row["amount"]),
        tax_amount=_clean_amount(row["tax_amount"]),
        is_excluded=bool(row["is_excluded"]),
        status=row["status"],
        engine_used=row["engine_used"],
        created_at=row["created_at"],
        file_size=file_size,
        has_source_file=has_source,
        source_url=source_url,
    )


def _get_conn():
    if db_manager is None:
        raise HTTPException(status_code=503, detail="database not initialized")
    return db_manager.get_connection()


# ---- Routes ----

@router.get("/health")
async def health_check():
    return {"status": "healthy", "service": "invoice"}


@router.post("/scan", response_model=Dict[str, Any])
async def scan_invoices(
    background_tasks: BackgroundTasks,
    files: List[UploadFile] = File(...),
):
    """
    Upload invoice PDF/image files and process them.

    Level 1: pdfplumber text extraction (searchable PDF)
    Level 2: Qwen2.5-VL vision OCR (scanned documents)

    Source files are ALWAYS preserved on disk (./data/invoices/), even when
    recognition fails, so users can re-process them later or compare against
    the extracted fields.
    """
    if not files:
        raise HTTPException(status_code=400, detail="no files provided")

    results = []
    processed_ids = []

    for file in files:
        # Save uploaded file (always, before processing, to preserve source for archival)
        safe_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
        file_path = INVOICE_DIR / safe_name
        content = await file.read()
        with open(file_path, "wb") as f:
            f.write(content)
        try:
            file_size = os.path.getsize(file_path)
        except OSError:
            file_size = 0

        # Process invoice
        from skills.invoice_skill import process_invoice, save_invoice_to_db
        try:
            result = await process_invoice(str(file_path), file.filename)
            conn = _get_conn()
            invoice_id = save_invoice_to_db(conn, result)
            processed_ids.append(invoice_id)
            results.append({
                "filename": file.filename,
                "stored_as": safe_name,
                "status": result.get("status"),
                "engine_used": result.get("engine_used"),
                "invoice_id": invoice_id,
            })
        except Exception as e:
            logger.error(f"[Invoice] Processing failed for {file.filename}: {e}")
            # Preserve the source file by recording a failed entry; user can still
            # download the original via /api/invoice/source/{id} and re-scan later.
            try:
                conn = _get_conn()
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO invoices
                        (filename, file_path, file_size, status, error_message,
                         engine_used, created_at, updated_at)
                    VALUES (?, ?, ?, 'failed', ?, NULL, ?, ?)
                """, (
                    file.filename,
                    str(file_path),
                    file_size,
                    str(e)[:500],
                    datetime.now().isoformat(),
                    datetime.now().isoformat(),
                ))
                conn.commit()
                failed_id = cursor.lastrowid
            except Exception as db_err:
                logger.error(f"[Invoice] Failed to record failed-scan entry: {db_err}")
                failed_id = None
            results.append({
                "filename": file.filename,
                "stored_as": safe_name,
                "status": "failed",
                "error": str(e),
                "invoice_id": failed_id,
            })

    return {
        "status": "ok",
        "processed": len(processed_ids),
        "total": len(files),
        "results": results,
    }


@router.get("/list", response_model=InvoiceListResponse)
async def list_invoices(
    request: Request,
    seller: Optional[str] = Query(None, description="Filter by seller name"),
    from_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    to_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    status: Optional[str] = Query(None, description="Filter by status"),
    excluded: Optional[bool] = Query(None, description="Filter by excluded status"),
    limit: int = Query(100, description="Max results"),
    offset: int = Query(0, description="Offset for pagination"),
):
    """List invoices with optional filters"""
    conn = _get_conn()
    cursor = conn.cursor()

    where = []
    params = []

    if seller:
        where.append("invoices.seller_name LIKE ?")
        params.append(f"%{seller}%")
    if from_date:
        where.append("invoices.invoice_date >= ?")
        params.append(from_date)
    if to_date:
        where.append("invoices.invoice_date <= ?")
        params.append(to_date)
    if status:
        where.append("invoices.status = ?")
        params.append(status)
    if excluded is not None:
        where.append("invoices.is_excluded = ?")
        params.append(1 if excluded else 0)

    where_clause = " AND ".join(where) if where else "1=1"

    cursor.execute(f"SELECT COUNT(*) FROM invoices WHERE {where_clause}", params)
    total = cursor.fetchone()[0]

    cursor.execute(
        f"SELECT * FROM invoices WHERE {where_clause} ORDER BY created_at DESC LIMIT ? OFFSET ?",
        params + [limit, offset],
    )
    base = str(request.base_url).rstrip("/")
    invoices = [_row_to_invoice(row, source_url_builder=lambda i, _b=base: f"{_b}/api/invoice/source/{i}") for row in cursor.fetchall()]

    return InvoiceListResponse(status="ok", count=total, invoices=invoices)


@router.get("/query", response_model=InvoiceListResponse)
async def query_invoices(
    request: Request,
    seller: Optional[str] = Query(None),
    buyer: Optional[str] = Query(None),
    from_date: Optional[str] = Query(None, alias="from"),
    to_date: Optional[str] = Query(None, alias="to"),
    invoice_number: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    limit: int = Query(100),
):
    """Query invoices by various criteria"""
    conn = _get_conn()
    cursor = conn.cursor()
    where = []
    params = []

    if seller:
        where.append("invoices.seller_name LIKE ?")
        params.append(f"%{seller}%")
    if buyer:
        where.append("invoices.buyer_name LIKE ?")
        params.append(f"%{buyer}%")
    if from_date:
        where.append("invoices.invoice_date >= ?")
        params.append(from_date)
    if to_date:
        where.append("invoices.invoice_date <= ?")
        params.append(to_date)
    if invoice_number:
        where.append("invoices.invoice_number = ?")
        params.append(invoice_number)
    if keyword:
        where.append("(invoices.seller_name LIKE ? OR invoices.buyer_name LIKE ? OR invoices.invoice_number LIKE ?)")
        kw = f"%{keyword}%"
        params.extend([kw, kw, kw])

    where_clause = " AND ".join(where) if where else "1=1"
    cursor.execute(f"SELECT COUNT(*) FROM invoices WHERE {where_clause}", params)
    total = cursor.fetchone()[0]
    cursor.execute(
        f"SELECT * FROM invoices WHERE {where_clause} ORDER BY invoice_date DESC LIMIT ?",
        params + [limit],
    )
    base = str(request.base_url).rstrip("/")
    invoices = [_row_to_invoice(row, source_url_builder=lambda i, _b=base: f"{_b}/api/invoice/source/{i}") for row in cursor.fetchall()]
    return InvoiceListResponse(status="ok", count=total, invoices=invoices)


@router.get("/detail/{invoice_id}")
async def get_invoice_detail(invoice_id: int, request: Request):
    """Get detailed invoice info including line items"""
    conn = _get_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM invoices WHERE id = ?", (invoice_id,))
    row = cursor.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="invoice not found")

    base = str(request.base_url).rstrip("/")
    invoice = _row_to_invoice(row, source_url_builder=lambda i, _b=base: f"{_b}/api/invoice/source/{i}")
    cursor.execute("SELECT * FROM invoice_items WHERE invoice_id = ?", (invoice_id,))
    items = [dict(item) for item in cursor.fetchall()]

    return {
        "status": "ok",
        "invoice": invoice.model_dump(),
        "items": items,
    }


@router.post("/exclude/{invoice_id}")
async def exclude_invoice(invoice_id: int):
    """Mark invoice as excluded from reimbursement"""
    conn = _get_conn()
    cursor = conn.cursor()
    cursor.execute("UPDATE invoices SET is_excluded = 1, updated_at = ? WHERE id = ?",
                   (datetime.now().isoformat(), invoice_id))
    conn.commit()
    if cursor.rowcount == 0:
        raise HTTPException(status_code=404, detail="invoice not found")
    return {"status": "ok", "message": f"Invoice #{invoice_id} excluded"}


@router.post("/include/{invoice_id}")
async def include_invoice(invoice_id: int):
    """Restore invoice for reimbursement"""
    conn = _get_conn()
    cursor = conn.cursor()
    cursor.execute("UPDATE invoices SET is_excluded = 0, updated_at = ? WHERE id = ?",
                   (datetime.now().isoformat(), invoice_id))
    conn.commit()
    if cursor.rowcount == 0:
        raise HTTPException(status_code=404, detail="invoice not found")
    return {"status": "ok", "message": f"Invoice #{invoice_id} included"}


@router.get("/export")
async def export_invoices(
    request: Request,
    from_date: Optional[str] = Query(None, alias="from"),
    to_date: Optional[str] = Query(None, alias="to"),
    format: str = Query("xlsx", description="xlsx or csv"),
    include_source_link: bool = Query(True, description="Append a HYPERLINK column to the original source file"),
):
    """Export invoices to Excel or CSV"""
    conn = _get_conn()
    cursor = conn.cursor()
    where = ["invoices.is_excluded = 0"]
    params = []
    if from_date:
        where.append("invoices.invoice_date >= ?")
        params.append(from_date)
    if to_date:
        where.append("invoices.invoice_date <= ?")
        params.append(to_date)
    where_clause = " AND ".join(where)
    cursor.execute(f"SELECT * FROM invoices WHERE {where_clause} ORDER BY invoice_date DESC", params)
    rows = cursor.fetchall()

    if not rows:
        raise HTTPException(status_code=404, detail="no invoices found for export")

    base_url = str(request.base_url).rstrip("/")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    source_link_written = 0

    if format == "csv":
        filename = f"invoices_export_{timestamp}.csv"
        output_path = OUTPUT_DIR / filename
        import csv
        with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            header = ["ID", "发票号码", "发票代码", "开票日期", "销售方", "销售方税号",
                      "购买方", "购买方税号", "金额", "税额", "价税合计", "状态", "识别引擎"]
            if include_source_link:
                header.append("源文件下载")
            writer.writerow(header)
            for r in rows:
                row_data = [
                    r["id"], r["invoice_number"], r["invoice_code"], r["invoice_date"],
                    r["seller_name"], r["seller_tax_id"], r["buyer_name"], r["buyer_tax_id"],
                    r["amount"], r["tax_amount"], r["total_amount"],
                    r["status"], r["engine_used"],
                ]
                if include_source_link:
                    fp = r["file_path"]
                    if fp and os.path.exists(fp):
                        row_data.append(f"{base_url}/api/invoice/source/{r['id']}")
                        source_link_written += 1
                    else:
                        row_data.append("")
                writer.writerow(row_data)
    else:
        filename = f"invoices_export_{timestamp}.xlsx"
        output_path = OUTPUT_DIR / filename
        import pandas as pd
        data = []
        for r in rows:
            data.append({
                "ID": r["id"],
                "发票号码": r["invoice_number"],
                "发票代码": r["invoice_code"],
                "开票日期": r["invoice_date"],
                "销售方": r["seller_name"],
                "销售方税号": r["seller_tax_id"],
                "购买方": r["buyer_name"],
                "购买方税号": r["buyer_tax_id"],
                "金额": r["amount"],
                "税额": r["tax_amount"],
                "价税合计": r["total_amount"],
                "状态": r["status"],
                "识别引擎": r["engine_used"],
            })
        df = pd.DataFrame(data)
        df.to_excel(output_path, index=False, engine="openpyxl")

        if include_source_link:
            try:
                id_to_path = {r["id"]: r["file_path"] for r in rows}
                source_link_written = _add_source_file_column(output_path, base_url, id_to_path)
            except Exception as e:
                logger.warning(f"[Invoice] Failed to add source file column: {e}")

    return {
        "status": "ok",
        "filename": filename,
        "path": str(output_path),
        "download_url": f"/api/excel/download/{filename}",
        "count": len(rows),
        "source_links": source_link_written,
    }


@router.get("/export-advanced")
async def export_invoices_advanced(
    request: Request,
    from_date: Optional[str] = Query(None, alias="from"),
    to_date: Optional[str] = Query(None, alias="to"),
    add_formulas: bool = Query(True, description="Add formula columns via minimax-xlsx"),
    run_validation: bool = Query(True, description="Run formula_check.py after export"),
    include_source_link: bool = Query(True, description="Append a HYPERLINK column to the original source file"),
):
    """
    Advanced Excel export using minimax-xlsx scripts (subprocess calls).

    Pipeline:
      1. openpyxl → base xlsx
      2. xlsx_unpack.py → unpack to temp dir
      3. xlsx_add_column.py → add formula column (校验: 金额+税额)
      4. formula_check.py → validate formulas
      5. xlsx_pack.py → repack final xlsx

    The 7B Agent only calls this one API; all complex CLI work is backend-managed.
    """
    SKILL_DIR = Path(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "skills", "minimax-xlsx", "scripts")))

    conn = _get_conn()
    cursor = conn.cursor()
    where = ["invoices.is_excluded = 0"]
    params = []
    if from_date:
        where.append("invoices.invoice_date >= ?")
        params.append(from_date)
    if to_date:
        where.append("invoices.invoice_date <= ?")
        params.append(to_date)
    where_clause = " AND ".join(where)
    cursor.execute(f"SELECT * FROM invoices WHERE {where_clause} ORDER BY invoice_date DESC", params)
    rows = cursor.fetchall()
    if not rows:
        raise HTTPException(status_code=404, detail="no invoices found for export")

    base_url = str(request.base_url).rstrip("/")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"invoices_advanced_{timestamp}"
    xlsx_path = OUTPUT_DIR / f"{base_name}.xlsx"

    # Step 1: Create base xlsx with openpyxl
    import pandas as pd
    data = []
    for r in rows:
        data.append({
            "ID": r["id"],
            "发票号码": r["invoice_number"],
            "发票代码": r["invoice_code"],
            "开票日期": r["invoice_date"],
            "销售方": r["seller_name"],
            "销售方税号": r["seller_tax_id"],
            "购买方": r["buyer_name"],
            "购买方税号": r["buyer_tax_id"],
            "金额": r["amount"],
            "税额": r["tax_amount"],
            "价税合计": r["total_amount"],
            "状态": r["status"],
            "识别引擎": r["engine_used"],
        })
    df = pd.DataFrame(data)
    df.to_excel(xlsx_path, index=False, engine="openpyxl")

    result = {
        "status": "ok",
        "filename": xlsx_path.name,
        "path": str(xlsx_path),
        "download_url": f"/api/excel/download/{xlsx_path.name}",
        "count": len(rows),
        "pipeline": [],
        "pipeline_ok": False,
    }

    if not add_formulas:
        if include_source_link:
            try:
                id_to_path = {r["id"]: r["file_path"] for r in rows}
                written = _add_source_file_column(xlsx_path, base_url, id_to_path, col_letter="O")
                result["source_links"] = written
            except Exception as e:
                logger.warning(f"[Invoice] Failed to add source-link column (no-formula): {e}")
                result["source_links"] = 0
        return result

    # Step 2: Unpack with xlsx_unpack.py
    import tempfile, subprocess
    work_dir = Path(tempfile.mkdtemp(prefix="invoice_xlsx_"))
    try:
        unpack_script = SKILL_DIR / "xlsx_unpack.py"
        if unpack_script.exists():
            r1 = subprocess.run(
                [sys.executable, str(unpack_script), str(xlsx_path), str(work_dir)],
                capture_output=True, text=True, timeout=30
            )
            result["pipeline"].append({
                "step": "unpack",
                "exit_code": r1.returncode,
                "stderr": r1.stderr[:200] if r1.returncode != 0 else "",
            })

            # Step 3: Add formula column using xlsx_add_column.py
            add_col_script = SKILL_DIR / "xlsx_add_column.py"
            if add_col_script.exists() and r1.returncode == 0:
                data_rows = len(rows)
                r2 = subprocess.run(
                    [sys.executable, str(add_col_script), str(work_dir),
                     "--col", "N",
                     "--sheet", "Sheet1",
                     "--header", "校验(金额+税额)",
                     "--formula", "=I{row}+J{row}",
                     "--formula-rows", f"2:{data_rows+1}",
                     "--numfmt", "#,##0.00",
                     "--border-row", str(data_rows + 1),
                     "--border-style", "medium"],
                    capture_output=True, text=True, timeout=30
                )
                result["pipeline"].append({
                    "step": "add_formula_column",
                    "exit_code": r2.returncode,
                    "stderr": r2.stderr[:200] if r2.returncode != 0 else "",
                })

                # Step 4: Run formula_check.py for validation
                if run_validation:
                    check_script = SKILL_DIR / "formula_check.py"
                    if check_script.exists():
                        r3 = subprocess.run(
                            [sys.executable, str(check_script), str(xlsx_path), "--summary"],
                            capture_output=True, text=True, timeout=30
                        )
                        result["pipeline"].append({
                            "step": "formula_validation",
                            "exit_code": r3.returncode,
                            "stdout": r3.stdout[:300],
                            "stderr": r3.stderr[:200] if r3.returncode != 0 else "",
                        })

            # Step 5: Pack with xlsx_pack.py
            pack_script = SKILL_DIR / "xlsx_pack.py"
            if pack_script.exists():
                r4 = subprocess.run(
                    [sys.executable, str(pack_script), str(work_dir), str(xlsx_path)],
                    capture_output=True, text=True, timeout=30
                )
                result["pipeline"].append({
                    "step": "pack",
                    "exit_code": r4.returncode,
                    "stderr": r4.stderr[:200] if r4.returncode != 0 else "",
                })
        else:
            result["pipeline"].append({"step": "unpack", "warning": "xlsx_unpack.py not found"})
    except subprocess.TimeoutExpired:
        result["pipeline"].append({"step": "error", "message": "minimax-xlsx subprocess timed out"})
    except Exception as e:
        result["pipeline"].append({"step": "error", "message": str(e)})
    finally:
        import shutil
        try:
            shutil.rmtree(str(work_dir), ignore_errors=True)
        except Exception:
            pass

    # Step 6: Append the source-file download HYPERLINK column (column O),
    # right after the validation column (N). Always runs — gives the auditor
    # a one-click way to compare an extracted row with the original scan.
    if include_source_link:
        try:
            id_to_path = {r["id"]: r["file_path"] for r in rows}
            written = _add_source_file_column(xlsx_path, base_url, id_to_path, col_letter="O")
            result["source_links"] = written
            result["pipeline"].append({
                "step": "add_source_link_column",
                "exit_code": 0,
                "rows_with_link": written,
            })
        except Exception as e:
            logger.warning(f"[Invoice] Failed to add source-link column (advanced): {e}")
            result["pipeline"].append({
                "step": "add_source_link_column",
                "exit_code": 1,
                "stderr": str(e)[:200],
            })

    result["pipeline_ok"] = all(
        p.get("exit_code", -1) == 0 for p in result["pipeline"]
    )
    return result


@router.get("/problems")
async def get_problem_invoices(request: Request):
    """List invoices with processing issues"""
    conn = _get_conn()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT * FROM invoices WHERE status = 'failed' OR (status = 'processed' AND "
        "(invoice_number IS NULL OR seller_name IS NULL OR total_amount IS NULL)) "
        "ORDER BY created_at DESC"
    )
    rows = cursor.fetchall()
    base = str(request.base_url).rstrip("/")
    invoices = [_row_to_invoice(row, source_url_builder=lambda i, _b=base: f"{_b}/api/invoice/source/{i}") for row in rows]
    return InvoiceListResponse(status="ok", count=len(invoices), invoices=invoices)


@router.get("/stats")
async def get_invoice_stats():
    """Get invoice statistics"""
    conn = _get_conn()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM invoices")
    total = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM invoices WHERE is_excluded = 0")
    active = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM invoices WHERE is_excluded = 1")
    excluded = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM invoices WHERE status = 'failed'")
    failed = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM invoices WHERE status = 'processed' AND engine_used = 'pdfplumber'")
    pdf_extracted = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM invoices WHERE status = 'processed' AND engine_used = 'qwen2.5vl'")
    vision_ocr = cursor.fetchone()[0]

    cursor.execute("""
        SELECT COALESCE(SUM(total_amount), 0) FROM invoices
        WHERE is_excluded = 0 AND total_amount IS NOT NULL
    """)
    total_amount = cursor.fetchone()[0]

    cursor.execute("""
        SELECT invoice_date, COUNT(*) as cnt FROM invoices
        WHERE invoice_date IS NOT NULL
        GROUP BY invoice_date ORDER BY invoice_date DESC LIMIT 10
    """)
    by_date = [{"date": r["invoice_date"], "count": r["cnt"]} for r in cursor.fetchall()]

    cursor.execute("""
        SELECT seller_name, COUNT(*) as cnt FROM invoices
        WHERE seller_name IS NOT NULL
        GROUP BY seller_name ORDER BY cnt DESC LIMIT 10
    """)
    top_sellers = [{"seller": r["seller_name"], "count": r["cnt"]} for r in cursor.fetchall()]

    return {
        "status": "ok",
        "stats": {
            "total": total,
            "active": active,
            "excluded": excluded,
            "failed": failed,
            "pdf_extracted": pdf_extracted,
            "vision_ocr": vision_ocr,
            "total_amount": total_amount,
        },
        "by_date": by_date,
        "top_sellers": top_sellers,
    }


@router.delete("/delete/{invoice_id}")
async def delete_invoice(invoice_id: int):
    """Delete an invoice record"""
    conn = _get_conn()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM invoices WHERE id = ?", (invoice_id,))
    conn.commit()
    if cursor.rowcount == 0:
        raise HTTPException(status_code=404, detail="invoice not found")
    return {"status": "ok", "message": f"Invoice #{invoice_id} deleted"}


# ---- Source file archival ----

_MIME_BY_EXT = {
    ".pdf": "application/pdf",
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".webp": "image/webp",
    ".bmp": "image/bmp",
    ".tiff": "image/tiff",
    ".tif": "image/tiff",
    ".ofd": "application/ofd",
}


def _media_type_for(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    return _MIME_BY_EXT.get(ext, "application/octet-stream")


def _add_source_file_column(
    xlsx_path: Path,
    base_url: str,
    id_to_path: Dict[int, str],
    *,
    col_letter: str = "O",
    col_header: str = "源文件下载",
) -> int:
    """
    Append a HYPERLINK column to an existing xlsx so each row links back to
    the original uploaded source file. Returns the number of rows that got
    a real link (rows with a missing source file are left blank).

    `id_to_path` maps invoice id -> file_path on disk; missing keys/paths are
    treated as "no archive copy" and the cell is left empty.
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb.active

    header_row = 1
    # Find the first data row (header is row 1, data starts at row 2)
    first_data_row = header_row + 1
    last_data_row = ws.max_row

    # Write header
    ws[f"{col_letter}{header_row}"] = col_header
    ws[f"{col_letter}{header_row}"].font = ws[f"A{header_row}"].font
    ws[f"{col_letter}{header_row}"].fill = ws[f"A{header_row}"].fill

    base = base_url.rstrip("/")
    written = 0
    for r in range(first_data_row, last_data_row + 1):
        id_cell = ws.cell(row=r, column=1).value
        try:
            inv_id = int(id_cell)
        except (TypeError, ValueError):
            continue
        file_path = id_to_path.get(inv_id)
        if not file_path or not os.path.exists(file_path):
            ws.cell(row=r, column=ws[f"{col_letter}1"].column, value="源文件缺失").font = ws.cell(row=r, column=1).font
            continue
        url = f"{base}/api/invoice/source/{inv_id}"
        display = _safe_filename(os.path.basename(file_path), fallback=f"invoice_{inv_id}")
        cell = ws.cell(row=r, column=ws[f"{col_letter}1"].column)
        cell.value = f'=HYPERLINK("{url}","下载源文件:{display}")'
        cell.hyperlink = url  # openpyxl also exposes a clickable link on hover
        cell.font = ws.cell(row=r, column=1).font.copy(color="0563C1", underline="single")
        written += 1

    # Adjust column width
    try:
        col_idx = ws[f"{col_letter}1"].column
        letter_dim = ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter]
        letter_dim.width = max(letter_dim.width or 0, 40)
    except Exception:
        pass

    wb.save(xlsx_path)
    return written


def _lookup_invoice_row(invoice_id: int):
    conn = _get_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM invoices WHERE id = ?", (invoice_id,))
    row = cursor.fetchone()
    return conn, row


@router.get("/source/{invoice_id}")
async def download_source_file(invoice_id: int):
    """
    Download the original uploaded invoice file for archival/comparison.

    The source file is preserved on disk in INVOICE_DIR whenever an invoice
    is uploaded (even when recognition fails), so users can compare it with
    the extracted fields in the Excel export.
    """
    conn, row = _lookup_invoice_row(invoice_id)
    try:
        if not row:
            raise HTTPException(status_code=404, detail="invoice not found")

        file_path = row["file_path"]
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(
                status_code=410,
                detail="source file no longer exists on disk (was it manually removed?)",
            )

        download_name = _safe_filename(row["filename"], fallback=f"invoice_{invoice_id}")
        return FileResponse(
            path=file_path,
            filename=download_name,
            media_type=_media_type_for(file_path),
        )
    finally:
        conn.close()


@router.get("/sources-archive")
async def download_sources_archive(
    seller: Optional[str] = Query(None, description="Filter by seller name"),
    from_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    to_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    status: Optional[str] = Query(None, description="Filter by status"),
    excluded: Optional[bool] = Query(None, description="Filter by excluded status"),
    include_missing: bool = Query(False, description="Include invoices whose source file is missing as a manifest.txt entry"),
):
    """
    Bundle all original invoice source files matching the filters into a single
    ZIP archive for archival. Useful for handing the raw scans to auditors or
    keeping an offline copy alongside the Excel export.
    """
    conn = _get_conn()
    cursor = conn.cursor()
    where = []
    params = []

    if seller:
        where.append("seller_name LIKE ?")
        params.append(f"%{seller}%")
    if from_date:
        where.append("invoice_date >= ?")
        params.append(from_date)
    if to_date:
        where.append("invoice_date <= ?")
        params.append(to_date)
    if status:
        where.append("status = ?")
        params.append(status)
    if excluded is not None:
        where.append("is_excluded = ?")
        params.append(1 if excluded else 0)

    where_clause = " AND ".join(where) if where else "1=1"
    cursor.execute(
        f"SELECT id, filename, file_path, invoice_number, invoice_date, seller_name, "
        f"total_amount, status, engine_used FROM invoices WHERE {where_clause} ORDER BY created_at DESC",
        params,
    )
    rows = cursor.fetchall()
    conn.close()

    if not rows:
        raise HTTPException(status_code=404, detail="no invoices match the filter")

    # Build ZIP in memory so we can stream it as a single response and avoid
    # leaving temp files behind.
    buf = io.BytesIO()
    manifest_lines = [
        "发票源文件留档清单 (Invoice Source File Archive)",
        f"生成时间: {datetime.now().isoformat()}",
        f"记录总数: {len(rows)}",
        f"筛选: seller={seller or '*'}, from={from_date or '*'}, to={to_date or '*'}, "
        f"status={status or '*'}, excluded={excluded}",
        "-" * 80,
        f"{'ID':<6}{'发票号码':<18}{'开票日期':<14}{'销售方':<30}{'价税合计':<14}{'引擎':<10}{'源文件':<10}{'文件名'}",
    ]
    missing_ids: List[int] = []
    seen_names: Dict[str, int] = {}

    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for r in rows:
            inv_id = r["id"]
            original = r["filename"] or f"invoice_{inv_id}"
            archive_name = _safe_filename(original, fallback=f"invoice_{inv_id}")
            archive_name = f"{inv_id:04d}_{archive_name}"
            # Disambiguate duplicate basenames inside the archive
            if archive_name in seen_names:
                seen_names[archive_name] += 1
                stem, ext = os.path.splitext(archive_name)
                archive_name = f"{stem}_{seen_names[archive_name]}{ext}"
            else:
                seen_names[archive_name] = 0

            file_path = r["file_path"]
            if file_path and os.path.exists(file_path):
                try:
                    zf.write(file_path, arcname=archive_name)
                    file_status = "OK"
                except OSError as e:
                    file_status = f"ERR({e})"
                    missing_ids.append(inv_id)
            else:
                file_status = "MISSING"
                missing_ids.append(inv_id)

            manifest_lines.append(
                f"{inv_id:<6}{str(r['invoice_number'] or ''):<18}"
                f"{str(r['invoice_date'] or ''):<14}"
                f"{str(r['seller_name'] or '')[:28]:<30}"
                f"{str(r['total_amount'] if r['total_amount'] is not None else ''):<14}"
                f"{str(r['engine_used'] or ''):<10}"
                f"{file_status:<10}{original}"
            )

        if missing_ids and include_missing:
            manifest_lines.append("-" * 80)
            manifest_lines.append(
                f"以下 {len(missing_ids)} 条记录的源文件已缺失,仅在清单中保留条目: "
                + ", ".join(str(i) for i in missing_ids)
            )
        elif missing_ids:
            manifest_lines.append(
                f"提示: {len(missing_ids)} 条记录源文件已缺失(未写入ZIP),如需清单仍记入请加 include_missing=true"
            )

        zf.writestr("manifest.txt", "\n".join(manifest_lines))

    buf.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_name = f"invoices_sources_{timestamp}.zip"
    return _serve_zip_buffer(buf, zip_name)


def _serve_zip_buffer(buf: io.BytesIO, filename: str):
    """
    Helper to stream an in-memory ZIP buffer through FastAPI as a downloadable
    attachment. Built in-memory so no temp files leak on disk.
    """
    from fastapi.responses import Response
    data = buf.getvalue()
    return Response(
        content=data,
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(data)),
        },
    )


# ---- GTD Task Integration ----

@router.post("/{invoice_id}/create-task")
async def create_invoice_task(invoice_id: int):
    """
    Create a GTD reimbursement task from an invoice.
    The task is linked to the invoice via source_type='invoice' / source_id.
    """
    conn = _get_conn()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM invoices WHERE id = ?", (invoice_id,))
    row = cursor.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="invoice not found")

    # Deduplicate: check if task already exists for this invoice
    cursor.execute(
        "SELECT id FROM gtd_tasks WHERE source_type = 'invoice' AND source_id = ?",
        (invoice_id,),
    )
    existing = cursor.fetchone()
    if existing:
        return {"status": "ok", "task_id": existing["id"], "message": "报销任务已存在，请勿重复创建"}

    title = f"报销发票 {row['invoice_number'] or '#'+str(invoice_id)} - {row['seller_name'] or '未知商家'}"

    amount = _clean_amount(row["total_amount"]) or 0
    priority = "high" if amount > 10000 else ("medium" if amount > 1000 else "low")

    description = (
        f"发票 #{invoice_id}\n"
        f"发票号码: {row['invoice_number'] or '-'}\n"
        f"发票代码: {row['invoice_code'] or '-'}\n"
        f"销售方: {row['seller_name'] or '-'}\n"
        f"购买方: {row['buyer_name'] or '-'}\n"
        f"金额: ¥{amount:.2f}\n"
        f"税额: ¥{(_clean_amount(row['tax_amount']) or 0):.2f}\n"
        f"价税合计: ¥{(_clean_amount(row['total_amount']) or 0):.2f}\n"
        f"开票日期: {row['invoice_date'] or '-'}\n"
        f"识别引擎: {row['engine_used'] or '-'}"
    )

    cursor.execute(
        """
        INSERT INTO gtd_tasks
            (title, description, category, priority, due_date,
             source_type, source_id, is_completed, created_at, updated_at)
        VALUES (?, ?, 'today', ?, ?, 'invoice', ?, 0, datetime('now'), datetime('now'))
        """,
        (title, description, priority, row["invoice_date"], invoice_id),
    )

    task_id = cursor.lastrowid
    conn.commit()

    return {
        "status": "ok",
        "task_id": task_id,
        "message": "报销任务已创建，请在「任务管理」中查看",
    }


@router.get("/{invoice_id}/tasks")
async def get_invoice_tasks(invoice_id: int):
    """Get GTD tasks linked to an invoice"""
    conn = _get_conn()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT id, title, description, category, priority, due_date,
               is_completed, completed_at, created_at, updated_at
        FROM gtd_tasks
        WHERE source_type = 'invoice' AND source_id = ?
        ORDER BY created_at DESC
        """,
        (invoice_id,),
    )
    tasks = [dict(r) for r in cursor.fetchall()]
    return {"status": "ok", "tasks": tasks}
