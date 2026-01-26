"""
Excel Exporter for Antinet
Export four-color cards and analysis reports to Excel format
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from datetime import datetime
from typing import List, Dict, Any, Optional
import pandas as pd


class AntinetExcelExporter:
    """
    Antinet Excel 导出器
    
    功能：
    - 导出四色卡片到 Excel
    - 生成结构化分析报告
    - 创建数据可视化图表
    - 支持多工作表
    """
    
    # 四色卡片颜色定义
    CARD_COLORS = {
        'fact': 'ADD8E6',      # 🔵 蓝色 - 事实
        'interpret': '90EE90',  # 🟢 绿色 - 解释
        'risk': 'FFFF99',       # 🟡 黄色 - 风险
        'action': 'FFB6C1'      # 🔴 红色 - 行动
    }
    
    CARD_NAMES = {
        'fact': '🔵 事实卡片',
        'interpret': '🟢 解释卡片',
        'risk': '🟡 风险卡片',
        'action': '🔴 行动建议'
    }
    
    def __init__(self):
        """初始化导出器"""
        self.wb = None
        self.current_sheet = None
    
    def create_workbook(self) -> Workbook:
        """创建新的工作簿"""
        self.wb = Workbook()
        # 删除默认的 Sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        return self.wb
    
    def add_overview_sheet(self, analysis_info: Dict[str, Any]):
        """
        添加报告概览工作表
        
        Args:
            analysis_info: 分析信息字典，包含：
                - title: 报告标题
                - date: 分析日期
                - data_source: 数据来源
                - card_counts: 卡片数量统计
                - summary: 报告摘要
        """
        ws = self.wb.create_sheet("📊 报告概览", 0)
        
        # 标题
        ws['A1'] = analysis_info.get('title', 'Antinet 智能分析报告')
        ws['A1'].font = Font(size=16, bold=True, color='1F4E78')
        ws.merge_cells('A1:D1')
        
        # 基本信息
        row = 3
        info_items = [
            ('分析日期', analysis_info.get('date', datetime.now().strftime('%Y-%m-%d'))),
            ('分析师', 'Antinet 智能知识管家'),
            ('分析架构', '8-Agent 协作系统'),
            ('数据来源', analysis_info.get('data_source', 'N/A')),
        ]
        
        for label, value in info_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # 卡片统计
        row += 1
        ws[f'A{row}'] = '四色卡片统计'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        card_counts = analysis_info.get('card_counts', {})
        for card_type, count in card_counts.items():
            ws[f'A{row}'] = self.CARD_NAMES.get(card_type, card_type)
            ws[f'B{row}'] = count
            # 添加颜色标记
            ws[f'A{row}'].fill = PatternFill(start_color=self.CARD_COLORS.get(card_type, 'FFFFFF'), 
                                             fill_type='solid')
            row += 1
        
        # 报告摘要
        row += 1
        ws[f'A{row}'] = '报告摘要'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        summary = analysis_info.get('summary', '无摘要信息')
        ws[f'A{row}'] = summary
        ws.merge_cells(f'A{row}:D{row+5}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        return ws
    
    def add_cards_sheet(self, card_type: str, cards: List[Dict[str, Any]]):
        """
        添加卡片工作表
        
        Args:
            card_type: 卡片类型 (fact/interpret/risk/action)
            cards: 卡片列表
        """
        sheet_name = self.CARD_NAMES.get(card_type, card_type)
        ws = self.wb.create_sheet(sheet_name)
        
        # 设置表头
        headers = ['卡片ID', '标题', '内容', '置信度', '创建时间', '标签']
        if card_type == 'risk':
            headers.insert(3, '风险等级')
        elif card_type == 'action':
            headers.insert(3, '优先级')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.CARD_COLORS.get(card_type, '4472C4'), 
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # 填充卡片数据
        for row, card in enumerate(cards, 2):
            ws.cell(row, 1, card.get('id', ''))
            ws.cell(row, 2, card.get('title', ''))
            ws.cell(row, 3, card.get('content', ''))
            
            col = 4
            if card_type == 'risk':
                ws.cell(row, col, card.get('risk_level', 'N/A'))
                col += 1
            elif card_type == 'action':
                ws.cell(row, col, card.get('priority', 'N/A'))
                col += 1
            
            ws.cell(row, col, card.get('confidence', 0.0))
            ws.cell(row, col + 1, card.get('created_at', ''))
            ws.cell(row, col + 2, ', '.join(card.get('tags', [])))
            
            # 设置行背景色（浅色）
            for c in range(1, len(headers) + 1):
                ws.cell(row, c).fill = PatternFill(start_color=self.CARD_COLORS.get(card_type, 'FFFFFF') + '40', 
                                                   fill_type='solid')
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 30
        
        # 自动筛选
        ws.auto_filter.ref = ws.dimensions
        
        return ws
    
    def add_data_sheet(self, sheet_name: str, data: pd.DataFrame):
        """
        添加数据工作表
        
        Args:
            sheet_name: 工作表名称
            data: pandas DataFrame
        """
        ws = self.wb.create_sheet(sheet_name)
        
        # 写入表头
        for col, column_name in enumerate(data.columns, 1):
            cell = ws.cell(1, col, column_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', fill_type='solid')
        
        # 写入数据
        for row_idx, row_data in enumerate(data.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row_idx, col_idx, value)
        
        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        return ws
    
    def add_chart_sheet(self, sheet_name: str, chart_data: Dict[str, Any]):
        """
        添加图表工作表
        
        Args:
            sheet_name: 工作表名称
            chart_data: 图表数据，包含：
                - type: 图表类型 (bar/line/pie)
                - title: 图表标题
                - data: 数据 (DataFrame)
                - x_col: X轴列名
                - y_cols: Y轴列名列表
        """
        ws = self.wb.create_sheet(sheet_name)
        
        # 先写入数据
        data = chart_data['data']
        for col, column_name in enumerate(data.columns, 1):
            ws.cell(1, col, column_name)
        
        for row_idx, row_data in enumerate(data.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row_idx, col_idx, value)
        
        # 创建图表
        chart_type = chart_data.get('type', 'bar')
        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'line':
            chart = LineChart()
        elif chart_type == 'pie':
            chart = PieChart()
        else:
            chart = BarChart()
        
        chart.title = chart_data.get('title', '数据图表')
        chart.style = 10
        
        # 设置数据范围
        data_rows = len(data) + 1
        y_cols = chart_data.get('y_cols', [data.columns[1]])
        
        for y_col in y_cols:
            col_idx = list(data.columns).index(y_col) + 1
            values = Reference(ws, min_col=col_idx, min_row=2, max_row=data_rows)
            chart.add_data(values, titles_from_data=False)
        
        # 设置分类轴
        x_col = chart_data.get('x_col', data.columns[0])
        x_col_idx = list(data.columns).index(x_col) + 1
        cats = Reference(ws, min_col=x_col_idx, min_row=2, max_row=data_rows)
        chart.set_categories(cats)
        
        # 添加图表到工作表
        ws.add_chart(chart, "E2")
        
        return ws
    
    def export_analysis_report(
        self, 
        output_path: str,
        analysis_info: Dict[str, Any],
        cards_by_type: Dict[str, List[Dict[str, Any]]],
        data_sheets: Optional[Dict[str, pd.DataFrame]] = None,
        charts: Optional[List[Dict[str, Any]]] = None
    ):
        """
        导出完整的分析报告
        
        Args:
            output_path: 输出文件路径
            analysis_info: 分析信息
            cards_by_type: 按类型分组的卡片字典
            data_sheets: 额外的数据工作表 (可选)
            charts: 图表列表 (可选)
        """
        # 创建工作簿
        self.create_workbook()
        
        # 添加概览工作表
        self.add_overview_sheet(analysis_info)
        
        # 添加四色卡片工作表
        for card_type in ['fact', 'interpret', 'risk', 'action']:
            cards = cards_by_type.get(card_type, [])
            if cards:
                self.add_cards_sheet(card_type, cards)
        
        # 添加数据工作表
        if data_sheets:
            for sheet_name, df in data_sheets.items():
                self.add_data_sheet(sheet_name, df)
        
        # 添加图表工作表
        if charts:
            for idx, chart_data in enumerate(charts, 1):
                sheet_name = chart_data.get('name', f'图表{idx}')
                self.add_chart_sheet(sheet_name, chart_data)
        
        # 保存文件
        self.wb.save(output_path)
        return output_path
    
    def export_cards_simple(
        self,
        output_path: str,
        cards: List[Dict[str, Any]],
        title: str = "Antinet 卡片导出"
    ):
        """
        简单导出卡片列表到单个工作表
        
        Args:
            output_path: 输出文件路径
            cards: 卡片列表
            title: 工作表标题
        """
        self.create_workbook()
        ws = self.wb.create_sheet(title, 0)
        
        # 表头
        headers = ['卡片ID', '类型', '标题', '内容', '置信度', '创建时间', '标签']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # 数据
        for row, card in enumerate(cards, 2):
            card_type = card.get('type', 'unknown')
            ws.cell(row, 1, card.get('id', ''))
            ws.cell(row, 2, self.CARD_NAMES.get(card_type, card_type))
            ws.cell(row, 3, card.get('title', ''))
            ws.cell(row, 4, card.get('content', ''))
            ws.cell(row, 5, card.get('confidence', 0.0))
            ws.cell(row, 6, card.get('created_at', ''))
            ws.cell(row, 7, ', '.join(card.get('tags', [])))
            
            # 根据类型设置背景色
            for c in range(1, len(headers) + 1):
                ws.cell(row, c).fill = PatternFill(
                    start_color=self.CARD_COLORS.get(card_type, 'FFFFFF') + '40', 
                    fill_type='solid'
                )
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 30
        
        self.wb.save(output_path)
        return output_path


# 便捷函数
def export_cards_to_excel(
    cards: List[Dict[str, Any]], 
    output_path: str,
    title: str = "Antinet 卡片导出"
) -> str:
    """
    快速导出卡片到 Excel
    
    Args:
        cards: 卡片列表
        output_path: 输出路径
        title: 标题
    
    Returns:
        输出文件路径
    """
    exporter = AntinetExcelExporter()
    return exporter.export_cards_simple(output_path, cards, title)


def export_analysis_to_excel(
    output_path: str,
    analysis_info: Dict[str, Any],
    cards_by_type: Dict[str, List[Dict[str, Any]]],
    data_sheets: Optional[Dict[str, pd.DataFrame]] = None,
    charts: Optional[List[Dict[str, Any]]] = None
) -> str:
    """
    导出完整分析报告到 Excel
    
    Args:
        output_path: 输出路径
        analysis_info: 分析信息
        cards_by_type: 按类型分组的卡片
        data_sheets: 数据工作表 (可选)
        charts: 图表 (可选)
    
    Returns:
        输出文件路径
    """
    exporter = AntinetExcelExporter()
    return exporter.export_analysis_report(
        output_path, 
        analysis_info, 
        cards_by_type, 
        data_sheets, 
        charts
    )
