#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

wb = Workbook()
ws = wb.active
ws.title = "测试用例"

# Headers
headers = ['用例ID', '模块', '功能', '测试步骤', '预期结果', '状态']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)

# Test data
test_cases = [
    ('TC-001', '卡片', '创建蓝色卡片', '新建→选择蓝色→填写内容→保存', '创建成功', '已测试'),
    ('TC-002', '卡片', '创建绿色卡片', '新建→选择绿色→填写内容→保存', '创建成功', '已测试'),
    ('TC-003', '卡片', '创建黄色卡片', '新建→选择黄色→填写内容→保存', '创建成功', '已测试'),
    ('TC-004', '卡片', '创建红色卡片', '新建→选择红色→填写内容→保存', '创建成功', '已测试'),
    ('TC-005', '卡片', '编辑卡片', '选择卡片→编辑→保存', '更新成功', '已测试'),
    ('TC-006', '卡片', '删除卡片', '选择卡片→删除→确认', '删除成功', '已测试'),
    ('TC-007', '卡片', '搜索卡片', '输入关键词→搜索', '返回匹配结果', '已测试'),
    ('TC-101', '智能体', '任务执行器', '提交任务→执行', '任务分解执行', '已测试'),
    ('TC-102', '智能体', '事实生成器', '输入查询→生成事实', '返回事实内容', '已测试'),
    ('TC-103', '智能体', '解释器', '输入问题→分析', '返回解释', '已测试'),
    ('TC-201', '对话', '发送消息', '输入问题→发送', '返回回答', '已测试'),
    ('TC-202', '对话', '多轮对话', '连续对话', '上下文保持', '已测试'),
    ('TC-301', '知识库', '导入PDF', '上传PDF→解析', '解析成功', '已测试'),
    ('TC-302', '知识库', '语义搜索', '输入查询→搜索', '返回结果', '已测试'),
    ('TC-401', 'GTD', '创建任务', '输入内容→保存', '创建成功', '已测试'),
    ('TC-402', 'GTD', '修改优先级', '选择→修改', '更新成功', '已测试'),
    ('TC-501', '文档', 'PDF分析', '上传→分析', '返回结果', '已测试'),
    ('TC-502', '文档', 'Excel分析', '上传→分析', '生成图表', '已测试'),
    ('TC-503', '文档', 'PPT生成', '选择内容→生成', '生成文件', '已测试'),
    ('TC-601', '视觉', '上传图片', '上传图片→提问', '返回分析', '已测试'),
    ('PT-001', '性能', 'NPU延迟', '发送请求→测量', '<500ms', '已测试'),
    ('ST-001', '安全', '数据本地', '检查目录', '本地存储', '已测试'),
]

for row, row_data in enumerate(test_cases, 2):
    for col, value in enumerate(row_data, 1):
        ws.cell(row=row, column=col, value=value)

ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 12

wb.save('output/docs/测试用例_Antinet.xlsx')
print('Created: output/docs/测试用例_Antinet.xlsx')
