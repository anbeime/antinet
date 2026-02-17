#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "功能清单"

# Headers
headers = ['序号', '模块', '功能名称', '功能描述', '优先级', '状态']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.font = Font(bold=True, color='FFFFFF')

# Data - Four Color Cards
data = [
    (1, '四色卡片', '卡片创建', '创建四色卡片（蓝/绿/黄/红）', 'P0', '已完成'),
    (2, '四色卡片', '卡片编辑', '编辑已有卡片内容', 'P0', '已完成'),
    (3, '四色卡片', '卡片删除', '删除不需要的卡片', 'P0', '已完成'),
    (4, '四色卡片', '卡片搜索', '关键词搜索和筛选', 'P0', '已完成'),
    (5, '四色卡片', '卡片导出', '导出为Word/PPT/PDF格式', 'P1', '已完成'),
    (6, '四色卡片', '卡片统计', '统计各类卡片数量', 'P2', '已完成'),
    # Agent System
    (7, '八大智能体', '任务执行器', '任务分解和执行调度', 'P0', '已完成'),
    (8, '八大智能体', '事实生成器', '生成事实性内容', 'P0', '已完成'),
    (9, '八大智能体', '解释器', '提供分析和解释', 'P0', '已完成'),
    (10, '八大智能体', '风险检测器', '识别潜在风险', 'P0', '已完成'),
    (11, '八大智能体', '卡片分类器', '内容分类', 'P0', '已完成'),
    (12, '八大智能体', '行动顾问', '提供行动建议', 'P0', '已完成'),
    (13, '八大智能体', '记忆管理者', '上下文和记忆管理', 'P0', '已完成'),
    (14, '八大智能体', '协调器', '协调各智能体工作', 'P0', '已完成'),
    # AI Chat
    (15, 'AI对话', '智能问答', '基于知识库的AI对话', 'P0', '已完成'),
    (16, 'AI对话', '多轮对话', '支持上下文的多轮对话', 'P0', '已完成'),
    (17, 'AI对话', '模型切换', '切换不同的AI模型', 'P1', '已完成'),
    (18, 'AI对话', '对话历史', '查看和管理对话历史', 'P1', '已完成'),
    # Knowledge Base
    (19, '知识库', '文档导入', '导入PDF/Word/Excel/PPT/TXT', 'P0', '已完成'),
    (20, '知识库', '智能解析', '自动解析文档提取关键信息', 'P0', '已完成'),
    (21, '知识库', '语义搜索', '基于向量的语义搜索', 'P0', '已完成'),
    (22, '知识库', '知识图谱', '可视化知识点关联', 'P1', '已完成'),
    (23, '知识库', '知识问答', '基于知识库的智能问答', 'P1', '已完成'),
    # GTD
    (24, 'GTD任务', '任务创建', '创建GTD任务', 'P0', '已完成'),
    (25, 'GTD任务', '任务分类', '分类管理', 'P0', '已完成'),
    (26, 'GTD任务', '优先级设置', '设置任务优先级', 'P1', '已完成'),
    (27, 'GTD任务', '截止日期', '设置截止日期和提醒', 'P1', '已完成'),
    (28, 'GTD任务', '任务统计', '任务完成情况统计', 'P2', '已完成'),
    # Document
    (29, '文档处理', 'PDF解析', '解析PDF文档提取文本', 'P0', '已完成'),
    (30, '文档处理', 'PDF分析', 'PDF内容智能分析', 'P1', '已完成'),
    (31, '文档处理', 'Excel分析', '数据分析与图表生成', 'P0', '已完成'),
    (32, '文档处理', 'PPT生成', '基于内容生成PPT', 'P0', '已完成'),
    (33, '文档处理', 'Word导出', '导出为Word文档', 'P1', '已完成'),
    # Vision
    (34, '视觉理解', '图像识别', '分析图片内容', 'P1', '已完成'),
    (35, '视觉理解', '图表解读', '理解数据图表', 'P1', '已完成'),
    (36, '视觉理解', '截图分析', '粘贴截图分析', 'P1', '已完成'),
    (37, '视觉理解', '多模态对话', '结合图像的对话', 'P1', '已完成'),
    # Skills
    (38, '技能系统', '技能中心', '技能管理和展示', 'P1', '已完成'),
    (39, '技能系统', '技能执行', '执行各类技能', 'P1', '已完成'),
    (40, '技能系统', '技能扩展', '支持自定义技能', 'P2', '进行中'),
]

for row, row_data in enumerate(data, 2):
    for col, value in enumerate(row_data, 1):
        ws.cell(row=row, column=col, value=value)

# Adjust column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 12

wb.save('output/docs/功能清单_Antinet.xlsx')
print('Created: output/docs/功能清单_Antinet.xlsx')
